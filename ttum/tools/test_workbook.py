"""End-to-end test: take the accounts, narration templates and settings out of the
built workbook, feed the bank's own figures through the record layout, and check
the result matches the file the bank sent.

This exercises the shipped workbook content, not just the layout constants.
"""

import os
import sys
from datetime import date, timedelta

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from ttum_reference import build_file, substitute_tokens  # noqa: E402

XLSM = os.path.join(HERE, "..", "dist", "TTUM_Generator.xlsm")
SAMPLE = os.path.join(HERE, "..", "samples", "PROPELG_TTUM_16072026_Revised.txt")

# The bank's figures for value date 15-07-2026, exactly as they would be typed
# into the Amount column under the default Paise setting: the last two digits are
# the paise, so 47400 is 474.00.
SAMPLE_AMOUNTS_KEYED = {
    "VI settlement": 20000,
    "MC settlement": 30000,
    "MC commission received": 1000,
    "VI commission received": 1000,
    "MC GST on commission": 200,
    "VI GST on commission": 200,
    "MC Non-GST commission": 100,
    "VI Non-GST commission": 100,
    "Net settlement to Prp India": 47400,
}

FIRST_ROW, LAST_ROW = 5, 104


def fail(message):
    print("FAIL: " + message)
    return 1


def main():
    wb = openpyxl.load_workbook(XLSM, keep_vba=True)
    entries, config = wb["Entries"], wb["Config"]

    settings = {config.cell(row=r, column=2).value: config.cell(row=r, column=3).value
                for r in range(5, 10)}

    # The macro divides the keyed figure by 100 unless Config says Rupees.
    unit = str(settings["Amount column is entered as"]).strip().lower()
    if unit not in ("paise", "rupees"):
        return fail("unexpected amount unit in Config: %r" % unit)
    divisor = 100.0 if unit == "paise" else 1.0

    value_date = date(2026, 7, 15)
    rows, total_dr, total_cr = [], 0.0, 0.0
    for r in range(FIRST_ROW, LAST_ROW + 1):
        include = entries.cell(row=r, column=1).value
        label = entries.cell(row=r, column=2).value
        if not include or str(include).strip().lower() != "yes":
            continue
        if label not in SAMPLE_AMOUNTS_KEYED:
            return fail("row %d is included but has no sample figure: %r" % (r, label))
        amount = SAMPLE_AMOUNTS_KEYED[label] / divisor
        drcr = entries.cell(row=r, column=4).value
        rows.append((entries.cell(row=r, column=3).value, amount, drcr,
                     entries.cell(row=r, column=6).value))
        if drcr == "D":
            total_dr += amount
        else:
            total_cr += amount

    if len(rows) != len(SAMPLE_AMOUNTS_KEYED):
        return fail("expected %d included rows, found %d"
                    % (len(SAMPLE_AMOUNTS_KEYED), len(rows)))
    if round(total_dr - total_cr, 2) != 0:
        return fail("the shipped rows do not balance: %.2f vs %.2f" % (total_dr, total_cr))

    trailing = str(settings["Write a line break after the last record"]).strip().lower() == "yes"
    produced = build_file(rows, value_date, trailing_newline=trailing)
    expected = open(SAMPLE, "rb").read()

    if produced != expected:
        got = produced.split(b"\r\n")
        want = expected.split(b"\r\n")
        for i, (a, b) in enumerate(zip(got, want), 1):
            if a != b:
                return fail("record %d differs:\n  got %r\n  exp %r" % (i, a, b))
        return fail("record count differs: %d vs %d" % (len(got), len(want)))

    # The file name the Dashboard formula and the macro will produce.
    offset = int(settings["File-name date offset (days)"])
    file_name = substitute_tokens(settings["File name pattern"], value_date + timedelta(days=offset))
    if file_name != "PROPELG_TTUM_16072026.txt":
        return fail("file name resolved to %r" % file_name)

    print("PASS: the workbook's own rows and settings reproduce the bank file exactly")
    print("      %d records, %d bytes, debit %.2f = credit %.2f, file name %s"
          % (len(rows), len(produced), total_dr, total_cr, file_name))
    print("      amounts keyed as %s: 47400 read back as %.2f" % (unit, 47400 / divisor))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
