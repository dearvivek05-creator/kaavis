"""End-to-end test of the import route.

Reads the real settlement file, matches its lines to the import keys held on the
built workbook's Entries sheet the same way the macro does, and then builds the
TTUM file from what the import would have loaded.
"""

import os
import sys
from datetime import date, timedelta

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from ttum_reference import build_file, substitute_tokens, RECORD_LEN, NARRATION_LEN  # noqa: E402

XLSM = os.path.join(HERE, "..", "dist", "TTUM_Generator.xlsm")
INPUT = os.path.join(HERE, "..", "samples", "NET_MERPAY_PROPELG__03092026032926.txt")

FIRST_ROW, LAST_ROW = 5, 104
LEN_ACCOUNT, LEN_DATE, LEN_AMOUNT = 14, 8, 23


def parse_line(raw):
    """Mirror of ParseInputLine in modTTUM.bas."""
    if len(raw) < 116:
        return None
    ddmmyyyy = raw[LEN_ACCOUNT:LEN_ACCOUNT + LEN_DATE]
    amount = raw[LEN_ACCOUNT + LEN_DATE:LEN_ACCOUNT + LEN_DATE + LEN_AMOUNT].strip()
    if not ddmmyyyy.isdigit() or not amount.isdigit():
        return None
    drcr = raw[LEN_ACCOUNT + LEN_DATE + LEN_AMOUNT]
    if drcr not in ("D", "C"):
        return None
    return dict(
        account=raw[:LEN_ACCOUNT].strip(),
        value_date=date(int(ddmmyyyy[4:8]), int(ddmmyyyy[2:4]), int(ddmmyyyy[0:2])),
        paise=int(amount),
        drcr=drcr,
        narration=raw[81:81 + 35].strip(),
    )


def fail(message):
    print("FAIL: " + message)
    return 1


def main():
    wb = openpyxl.load_workbook(XLSM, keep_vba=True)
    entries = wb["Entries"]
    config = {wb["Config"].cell(row=r, column=2).value: wb["Config"].cell(row=r, column=3).value
              for r in range(5, 13)}

    # The rows that can take an imported figure.
    keyed = []
    for r in range(FIRST_ROW, LAST_ROW + 1):
        key = entries.cell(row=r, column=8).value
        if key and str(key).strip():
            keyed.append((r, str(key).strip(),
                          entries.cell(row=r, column=2).value,
                          entries.cell(row=r, column=3).value,
                          entries.cell(row=r, column=6).value))
    if not keyed:
        return fail("no rows on the Entries sheet carry an import match text")

    raw_lines = open(INPUT, "rb").read().decode("ascii").split("\r\n")
    parsed = [p for p in (parse_line(l) for l in raw_lines if l.strip()) if p]
    if len(parsed) != 9:
        return fail("expected 9 usable records, parsed %d" % len(parsed))

    # Match exactly as MatchInputLines does: the row's key inside the narration.
    loaded, unmatched_lines = {}, []
    for line in parsed:
        hits = [k for k in keyed if k[1].lower() in line["narration"].lower()]
        if len(hits) != 1:
            unmatched_lines.append((line["narration"], len(hits)))
            continue
        row = hits[0][0]
        if row in loaded:
            return fail("two lines both claim Entries row %d" % row)
        loaded[row] = line

    if unmatched_lines:
        return fail("lines that did not match exactly one row: %r" % unmatched_lines)
    if len(loaded) != len(keyed):
        missing = [k[2] for k in keyed if k[0] not in loaded]
        return fail("rows left without a figure: %r" % missing)

    total_dr = sum(l["paise"] for l in loaded.values() if l["drcr"] == "D")
    total_cr = sum(l["paise"] for l in loaded.values() if l["drcr"] == "C")
    if total_dr != total_cr:
        return fail("imported figures do not balance: %d vs %d" % (total_dr, total_cr))

    # Now build the TTUM file from what the import loaded.
    value_date = parsed[0]["value_date"]
    rows = []
    for row, _key, _label, account, narration in keyed:
        line = loaded[row]
        rows.append((account, line["paise"] / 100.0, line["drcr"], narration))

    trailing = str(config["Write a line break after the last record"]).strip().lower() == "yes"
    produced = build_file(rows, value_date, trailing_newline=trailing).decode("ascii")
    records = produced.split("\r\n")

    if len(records) != len(keyed):
        return fail("produced %d records from %d rows" % (len(records), len(keyed)))
    for i, record in enumerate(records, 1):
        if len(record) != RECORD_LEN:
            return fail("record %d is %d characters" % (i, len(record)))
        if record[LEN_ACCOUNT:LEN_ACCOUNT + LEN_DATE] != value_date.strftime("%d%m%Y"):
            return fail("record %d carries the wrong value date" % i)

    # Every narration must still fit once the date is substituted in.
    for row, _key, label, _account, narration in keyed:
        text = substitute_tokens(narration, value_date)
        if len(text) > NARRATION_LEN:
            return fail("narration for %r is %d characters: %r" % (label, len(text), text))

    offset = int(config["File-name date offset (days)"])
    file_name = substitute_tokens(config["File name pattern"], value_date + timedelta(days=offset))

    print("PASS: the settlement file imports cleanly and generates a valid TTUM file")
    print("      %d records read, all %d matched 1:1 to the Entries sheet"
          % (len(parsed), len(loaded)))
    print("      value date %s, debit %.2f = credit %.2f"
          % (value_date.strftime("%d-%b-%Y"), total_dr / 100.0, total_cr / 100.0))
    print("      would produce %s, %d records of %d characters"
          % (file_name, len(records), RECORD_LEN))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
