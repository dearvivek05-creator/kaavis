"""Checks the macro-free Text Output sheet by actually evaluating its formulas.

The workbook is filled with the bank's own figures and then calculated with an
Excel formula interpreter, so the records this sheet produces are compared against
the bank's file the same way the macro's output is.
"""

import os
import sys
import warnings
from datetime import date

import openpyxl

warnings.filterwarnings("ignore")

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from test_workbook import SAMPLE_AMOUNTS_KEYED  # noqa: E402

XLSX = os.path.join(HERE, "..", "dist", "TTUM_Generator_NoMacros.xlsx")
SAMPLE = os.path.join(HERE, "..", "samples", "PROPELG_TTUM_16072026_Revised.txt")
SETTLEMENT = os.path.join(HERE, "..", "samples", "NET_MERPAY_PROPELG__03092026032926.txt")
IMPORT_FIRST = 10
SCRATCH = os.path.join(HERE, "..", "dist", "_formula_check.xlsx")

FIRST_ROW, LAST_ROW = 5, 104
OUT_FIRST = 9


def main():
    wb = openpyxl.load_workbook(XLSX)
    entries = wb["Entries"]

    # Point the workbook at the day the bank's sample covers, the way the
    # "generate for another date" path does: today off, an explicit date on.
    wb["Dashboard"]["C6"] = "No"
    wb["Dashboard"]["C7"] = date(2026, 7, 15)
    filled = 0
    for r in range(FIRST_ROW, LAST_ROW + 1):
        label = entries.cell(row=r, column=2).value
        if label in SAMPLE_AMOUNTS_KEYED:
            entries.cell(row=r, column=5).value = SAMPLE_AMOUNTS_KEYED[label]
            filled += 1
    if filled != len(SAMPLE_AMOUNTS_KEYED):
        print("FAIL: filled %d of %d rows" % (filled, len(SAMPLE_AMOUNTS_KEYED)))
        return 1

    # Paste a real settlement file into the Import sheet, the way someone would
    # when macros are blocked, so its parsing formulas are checked too.
    settlement = [l for l in open(SETTLEMENT, "rb").read().decode("ascii").split("\r\n")
                  if l.strip()]
    for i, line in enumerate(settlement):
        wb["Import"].cell(row=IMPORT_FIRST + i, column=2).value = line
    wb.save(SCRATCH)

    import formulas
    model = formulas.ExcelModel().loads(SCRATCH).finish()
    solution = model.calculate()
    os.remove(SCRATCH)

    book = os.path.basename(SCRATCH)
    records = []
    for i in range(len(SAMPLE_AMOUNTS_KEYED)):
        key = "'[%s]TEXT OUTPUT'!A%d" % (book, OUT_FIRST + i)
        cell = solution.get(key)
        if cell is None:
            print("FAIL: no calculated value for %s" % key)
            return 1
        value = cell.value[0, 0]
        records.append(str(value))

    # The Import sheet must read the settlement file back correctly.
    for i, line in enumerate(settlement):
        r = IMPORT_FIRST + i
        got = {col: solution.get("'[%s]IMPORT'!%s%d" % (book, col, r)) for col in "CEFGI"}
        if any(v is None for v in got.values()):
            print("FAIL: Import row %d did not calculate" % r)
            return 1
        account = str(got["C"].value[0, 0])
        paise = got["E"].value[0, 0]
        drcr = str(got["F"].value[0, 0])
        narration = str(got["G"].value[0, 0])
        goes_to = str(got["I"].value[0, 0])
        if account != line[:14].strip():
            print("FAIL: Import row %d account %r" % (r, account))
            return 1
        if int(paise) != int(line[22:45]):
            print("FAIL: Import row %d amount %r" % (r, paise))
            return 1
        if drcr != line[45] or narration != line[81:116].strip():
            print("FAIL: Import row %d parsed as %r / %r" % (r, drcr, narration))
            return 1
        if goes_to in ("-- no match --", ""):
            print("FAIL: Import row %d (%r) matched no Entries row" % (r, narration))
            return 1
    print("      Import sheet parsed all %d settlement lines and matched every one"
          % len(settlement))

    expected = open(SAMPLE, "rb").read().decode("ascii").split("\r\n")
    for i, (got, want) in enumerate(zip(records, expected), 1):
        if got != want:
            print("FAIL: record %d differs" % i)
            print("  got %r" % got)
            print("  exp %r" % want)
            return 1

    # The other branch of the date switch: "use today" must resolve to today.
    wb2 = openpyxl.load_workbook(XLSX)
    wb2["Dashboard"]["C6"] = "Yes"
    wb2["Dashboard"]["C7"] = None
    e2 = wb2["Entries"]
    for r in range(FIRST_ROW, LAST_ROW + 1):
        label = e2.cell(row=r, column=2).value
        if label in SAMPLE_AMOUNTS_KEYED:
            e2.cell(row=r, column=5).value = SAMPLE_AMOUNTS_KEYED[label]
    wb2.save(SCRATCH)
    today_sol = formulas.ExcelModel().loads(SCRATCH).finish().calculate()
    os.remove(SCRATCH)
    first = str(today_sol.get("'[%s]TEXT OUTPUT'!A%d" % (book, OUT_FIRST)).value[0, 0])
    stamped = first[14:22]
    wanted = date.today().strftime("%d%m%Y")
    if stamped != wanted:
        print("FAIL: with today selected the record carries %r, expected %r" % (stamped, wanted))
        return 1
    print("      with \"use today\" selected, records carry %s" % wanted)

    print("PASS: the Text Output formulas reproduce all %d records of the bank file"
          % len(records))
    lengths = set(len(r) for r in records)
    print("      every record is %d characters, as the bank expects" % lengths.pop())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
