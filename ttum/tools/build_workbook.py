#!/usr/bin/env python3
"""Builds TTUM_Generator.xlsm - the workbook plus its embedded VBA project.

Run:  python3 build_workbook.py
Out:  ../dist/TTUM_Generator.xlsm
      ../dist/TTUM_Generator_NoMacros.xlsx  (same sheets, for the manual-import route)
"""

import os
import re
import shutil
import sys
import zipfile

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import vbaproject  # noqa: E402
from standard_rows import STANDARD_ROWS  # noqa: E402

DIST = os.path.join(HERE, "..", "dist")
VBA = os.path.join(HERE, "..", "vba")

FIRST_ROW, LAST_ROW = 5, 104

# ---------------------------------------------------------------- appearance
INK = "1F3355"
MUTED = "6B7280"
NAVY = PatternFill("solid", fgColor="1F3355")
BAND = PatternFill("solid", fgColor="EEF2F7")
INPUT_FILL = PatternFill("solid", fgColor="FFF7DB")
HEAD_FILL = PatternFill("solid", fgColor="D9E2EC")
WHITE = PatternFill("solid", fgColor="FFFFFF")

TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
SUBTITLE = Font(name="Calibri", size=10, italic=True, color="D6E0EA")
SECTION = Font(name="Calibri", size=11, bold=True, color=INK)
LABEL = Font(name="Calibri", size=11, color="1F2937")
NOTE = Font(name="Calibri", size=9, italic=True, color=MUTED)
VALUE = Font(name="Calibri", size=11, bold=True, color="1F2937")
HEADER = Font(name="Calibri", size=10, bold=True, color=INK)
MONO = Font(name="Consolas", size=10)

THIN = Side(style="thin", color="BFCBD9")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def banner(ws, title, subtitle, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    for row in (1, 2):
        for col in range(1, last_col + 1):
            ws.cell(row=row, column=col).fill = NAVY
    c = ws.cell(row=1, column=1)
    c.value = "   " + title
    c.font = TITLE
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    c = ws.cell(row=2, column=1)
    c.value = "   " + subtitle
    c.font = SUBTITLE
    ws.row_dimensions[2].height = 16


def input_cell(ws, addr, value=None, fmt=None):
    c = ws[addr]
    if value is not None:
        c.value = value
    c.fill = INPUT_FILL
    c.font = VALUE
    c.border = BOX
    if fmt:
        c.number_format = fmt
    return c


# ------------------------------------------------------------------ sheets

def build_dashboard(ws):
    banner(ws, "TTUM GENERATOR",
           "Enter the day's amounts on the Entries sheet, check the date, then click Generate.", 8)
    widths = {"A": 2.5, "B": 34, "C": 46, "D": 2.5, "E": 2.5, "F": 2.5, "G": 2.5, "H": 30}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("B3:H3")
    ws["B3"] = ("Buttons doing nothing?  Close Excel, right-click this file in File Explorer, "
                "Properties, tick Unblock, reopen.  Full list on the Setup sheet.")
    ws["B3"].fill = PatternFill("solid", fgColor="FFF3CD")
    ws["B3"].font = Font(name="Calibri", size=10, bold=True, color="8A6100")
    ws["B3"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[3].height = 20

    def section(row, text):
        ws.cell(row=row, column=2, value=text).font = SECTION
        for col in (2, 3):
            ws.cell(row=row, column=col).fill = BAND

    def label(row, text):
        ws.cell(row=row, column=2, value=text).font = LABEL

    def note(row, text):
        c = ws.cell(row=row, column=2, value=text)
        c.font = NOTE
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

    section(5, "STEP 1   Set the date")
    label(6, "Value date  (goes into every record)")
    input_cell(ws, "C6", "=TODAY()", "dd-mmm-yyyy")
    label(7, "File-name date  (offset is set on Config)")
    input_cell(ws, "C7", "=ttValueDate+ttFileDateOffset", "dd-mmm-yyyy")
    note(8, "Today's date is filled in automatically. Type over it to build the file for another day.")

    section(10, "STEP 2   Enter the amounts, or import them")
    label(11, "Rows included")
    ws["C11"] = '=COUNTIF(Entries!$A$%d:$A$%d,"Yes")' % (FIRST_ROW, LAST_ROW)
    label(12, "Total debit")
    ws["C12"] = "=Entries!$L$1"
    label(13, "Total credit")
    ws["C13"] = "=Entries!$L$2"
    label(14, "Difference  (must be 0.00)")
    ws["C14"] = "=Entries!$L$3"
    for addr in ("C12", "C13", "C14"):
        ws[addr].number_format = "#,##0.00"
    for addr in ("C11", "C12", "C13", "C14"):
        ws[addr].font = VALUE
        ws[addr].border = BOX
    ws.conditional_formatting.add("C14", CellIsRule(
        operator="notEqual", formula=["0"],
        fill=PatternFill("solid", fgColor="FFC7CE"), font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add("C14", CellIsRule(
        operator="equal", formula=["0"],
        fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(color="006100", bold=True)))

    section(16, "STEP 3   Choose where the file goes, then generate")
    label(17, "Output folder")
    input_cell(ws, "C17")
    ws["C17"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[17].height = 20
    note(18, "Type or paste a full folder path here, for example  D:\\TTUM\\Daily  -  or click "
             "Choose Output Folder. Left blank, files go to a TTUM_Output folder beside this workbook.")
    ws.row_dimensions[18].height = 26
    ws["B18"].alignment = Alignment(wrap_text=True, vertical="top")
    label(19, "File name")
    ws["C19"] = ('=SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(ttFileNamePattern,'
                 '"{DDMMYYYY}",TEXT(ttFileDate,"ddmmyyyy")),'
                 '"{DDMMYY}",TEXT(ttFileDate,"ddmmyy")),'
                 '"{YYYY}",TEXT(ttFileDate,"yyyy"))')
    ws["C19"].font = VALUE
    ws["C19"].border = BOX

    note(20, "Type the amounts on the Entries sheet, or click Import Latest Input File to read "
             "them out of the settlement file. The input folder is set on the Config sheet.")
    ws.row_dimensions[20].height = 26
    ws["B20"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.cell(row=21, column=2, value="Status").font = SECTION
    ws.merge_cells("B22:H22")
    ws["B22"] = "Ready."
    ws["B22"].font = Font(name="Calibri", size=11, bold=True, color="006E3C")
    ws["B22"].alignment = Alignment(vertical="center")
    ws.row_dimensions[22].height = 22

    ws["B24"] = ("Every button is also on the Alt+F8 macro list, and the Text Output sheet "
                 "builds the same file with no macros at all.")
    ws["B24"].font = NOTE
    ws.merge_cells("B24:H24")

    ws.sheet_view.showGridLines = False


def build_entries(ws):
    banner(ws, "DAILY TTUM ENTRIES",
           "Type today's amount and Dr/Cr against each line. Blank rows are ignored.", 7)
    for col, w in {"A": 10, "B": 28, "C": 18, "D": 8, "E": 20, "F": 42,
                   "G": 16, "H": 26, "I": 6, "J": 2.5, "K": 14, "L": 14}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A3:H3")
    ws["A3"] = ('="Total debit  "&TEXT($L$1,"#,##0.00")&"      Total credit  "'
                '&TEXT($L$2,"#,##0.00")&"      Difference  "&TEXT($L$3,"#,##0.00")'
                '&IF($L$3=0,"    (balanced)","    <<< THESE MUST MATCH")')
    ws["A3"].font = Font(name="Calibri", size=11, bold=True)
    ws["A3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 20
    ws.conditional_formatting.add("A3:H3", FormulaRule(
        formula=["$L$3<>0"], fill=PatternFill("solid", fgColor="FFC7CE"),
        font=Font(color="9C0006", bold=True)))
    ws.conditional_formatting.add("A3:H3", FormulaRule(
        formula=["$L$3=0"], fill=PatternFill("solid", fgColor="C6EFCE"),
        font=Font(color="006100", bold=True)))

    divisor = 'IF(ttAmountUnit="Paise",100,1)'
    ws["K1"] = "Total debit"
    ws["L1"] = '=SUMIFS($E$%d:$E$%d,$D$%d:$D$%d,"D",$A$%d:$A$%d,"Yes")/%s' % (
        FIRST_ROW, LAST_ROW, FIRST_ROW, LAST_ROW, FIRST_ROW, LAST_ROW, divisor)
    ws["K2"] = "Total credit"
    ws["L2"] = '=SUMIFS($E$%d:$E$%d,$D$%d:$D$%d,"C",$A$%d:$A$%d,"Yes")/%s' % (
        FIRST_ROW, LAST_ROW, FIRST_ROW, LAST_ROW, FIRST_ROW, LAST_ROW, divisor)
    ws["K3"] = "Difference"
    ws["L3"] = "=$L$1-$L$2"
    for row in (1, 2, 3):
        ws.cell(row=row, column=11).font = NOTE
        ws.cell(row=row, column=12).number_format = "#,##0.00"
        ws.cell(row=row, column=12).font = Font(size=10, bold=True)

    headers = ["Include", "Description", "Account Number", "Dr/Cr",
               "Amount as keyed", "Narration template", "Reads as (INR)",
               "Import match text", "Seq"]
    for i, text in enumerate(headers, start=1):
        c = ws.cell(row=FIRST_ROW - 1, column=i, value=text)
        c.font = HEADER
        c.fill = HEAD_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[FIRST_ROW - 1].height = 28

    for offset, row in enumerate(STANDARD_ROWS):
        r = FIRST_ROW + offset
        ws.cell(row=r, column=1, value="Yes")
        ws.cell(row=r, column=2, value=row["label"])
        ws.cell(row=r, column=3, value=row["account"])
        ws.cell(row=r, column=4, value=row["type"])
        ws.cell(row=r, column=6, value=row["narration"])
        ws.cell(row=r, column=8, value=row["key"])

    for r in range(FIRST_ROW, LAST_ROW + 1):
        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.border = BOX
            c.fill = WHITE
            c.font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).number_format = "@"
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=5).number_format = "#,##0"
        ws.cell(row=r, column=5).fill = INPUT_FILL
        # A live read-back of the keyed figure, so the decimal point is never in doubt.
        ws.cell(row=r, column=7,
                value='=IF($E%d="","",$E%d/%s)' % (r, r, divisor))
        ws.cell(row=r, column=7).number_format = "#,##0.00"
        ws.cell(row=r, column=7).font = Font(name="Calibri", size=10, bold=True, color="1F3355")
        ws.cell(row=r, column=8).alignment = Alignment(vertical="center")
        # Position of this row among the included ones, so the Text Output sheet
        # can list the records with no gaps and without any macro.
        ws.cell(row=r, column=9,
                value='=IF(AND($A%d="Yes",$E%d<>""),'
                      'SUMPRODUCT(($A$%d:$A%d="Yes")*($E$%d:$E%d<>"")),"")'
                      % (r, r, FIRST_ROW, r, FIRST_ROW, r))
        ws.cell(row=r, column=9).font = Font(name="Calibri", size=9, color="9AA5B1")
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="center")

    inc = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True,
                         showErrorMessage=True, errorTitle="Include?",
                         error="Choose Yes to put this row in today's file, or No to leave it out.")
    drcr = DataValidation(type="list", formula1='"D,C"', allow_blank=True,
                          showErrorMessage=True, errorTitle="Dr/Cr",
                          error="Enter D for a debit or C for a credit.")
    amt = DataValidation(type="decimal", operator="greaterThan", formula1="0",
                         allow_blank=True, showErrorMessage=True, errorTitle="Amount",
                         error="Enter the amount as it appears on the settlement report.")
    for dv, col in ((inc, "A"), (drcr, "D"), (amt, "E")):
        ws.add_data_validation(dv)
        dv.add("%s%d:%s%d" % (col, FIRST_ROW, col, LAST_ROW))

    ws.freeze_panes = "A%d" % FIRST_ROW
    ws.sheet_view.showGridLines = False

    note_row = LAST_ROW + 2
    ws.cell(row=note_row, column=1,
            value="Amount as keyed: with Config set to Paise (the default), type the figure exactly as "
                  "the settlement report shows it - the last two digits are the paise, so 47400 is "
                  "474.00. The Reads as (INR) column shows how each figure will be understood. "
                  "Narration tokens {DDMMYY} {DDMMYYYY} {DD} {MM} {YY} {YYYY} are replaced with the "
                  "value date; the narration must fit 35 characters after substitution. "
                  "Import match text is what an incoming settlement line's narration must contain "
                  "for its amount to land on this row - only the amount and Dr/Cr are imported, "
                  "never the account number or the narration.").font = NOTE
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 1, end_column=8)
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")


def build_config(ws):
    banner(ws, "CONFIG", "Settings that rarely change. The daily inputs live on the Dashboard.", 4)
    for col, w in {"A": 2.5, "B": 36, "C": 32, "D": 62}.items():
        ws.column_dimensions[col].width = w

    rows = [
        ("Amount column is entered as", "Paise",
         "Paise: the Amount column is the figure exactly as the settlement report shows it, "
         "with the last two digits being the paise, so 47400 means 474.00. "
         "Rupees: type 474.00 instead. The Reads as (INR) column on Entries shows the effect."),
        ("File name pattern", "PROPELG_TTUM_{DDMMYYYY}.txt",
         "Tokens are replaced with the file-name date shown on the Dashboard."),
        ("File-name date offset (days)", 1,
         "The bank sample is dated one day after the value date it carries, so this is 1. "
         "Set 0 to use the value date itself."),
        ("Block generation when out of balance", "Yes",
         "Yes stops the file when total debit does not equal total credit. "
         "No asks first and lets you continue."),
        ("Write a line break after the last record", "No",
         "No matches the bank's sample file, which ends immediately after the last record."),
        ("Input folder", "",
         "The folder the settlement system drops its file into. Import Latest Input File "
         "takes the newest matching file from here. Leave blank to be asked each time."),
        ("Input file name pattern", "NET_MERPAY_PROPELG*.txt",
         "Which files in that folder count as settlement files. * stands for anything."),
        ("Imported file sets the value date", "Yes",
         "Yes takes the value date from the file's own records. No keeps whatever is "
         "already on the Dashboard."),
    ]
    ws.cell(row=4, column=2, value="Setting").font = HEADER
    ws.cell(row=4, column=3, value="Value").font = HEADER
    ws.cell(row=4, column=4, value="What it does").font = HEADER
    for col in (2, 3, 4):
        ws.cell(row=4, column=col).fill = HEAD_FILL
        ws.cell(row=4, column=col).border = BOX

    for i, (name, value, note) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=name).font = LABEL
        c = ws.cell(row=r, column=3, value=value)
        c.fill = INPUT_FILL
        c.font = VALUE
        c.border = BOX
        n = ws.cell(row=r, column=4, value=note)
        n.font = NOTE
        n.alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=2).border = BOX
        ws.row_dimensions[r].height = 44 if i == 0 else 30

    unit = DataValidation(type="list", formula1='"Paise,Rupees"', allow_blank=False)
    ws.add_data_validation(unit)
    unit.add("C5")
    yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(yn)
    yn.add("C8:C9")
    yn2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(yn2)
    yn2.add("C12")
    ws.sheet_view.showGridLines = False


def build_log(ws):
    banner(ws, "GENERATION LOG", "One row is added every time a file is written.", 10)
    headers = ["Generated at", "Value date", "File date", "File name", "Records",
               "Total debit", "Total credit", "Balance", "User", "Full path"]
    widths = [19, 14, 14, 30, 9, 14, 14, 16, 18, 60]
    for i, (text, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=3, column=i, value=text)
        c.font = HEADER
        c.fill = HEAD_FILL
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False


def build_layout(ws):
    banner(ws, "RECORD LAYOUT", "Reference only - the macro builds every record to this layout.", 6)
    for col, w in {"A": 2.5, "B": 8, "C": 14, "D": 24, "E": 10, "F": 60}.items():
        ws.column_dimensions[col].width = w

    headers = ["#", "Columns", "Field", "Length", "Content"]
    for i, text in enumerate(headers, start=2):
        c = ws.cell(row=4, column=i, value=text)
        c.font = HEADER
        c.fill = HEAD_FILL
        c.border = BOX

    fields = [
        ("1-14", "Account number", 14, "Left justified, padded on the right with spaces."),
        ("15-22", "Value date", 8, "DDMMYYYY."),
        ("23-45", "Amount", 23, "Whole paise, padded on the left with zeros. "
                                "20000 means 200.00."),
        ("46", "Transaction type", 1, "D for debit, C for credit."),
        ("47-55", "Value date", 9, "The same date again as 0 + DDMMYYYY."),
        ("56-81", "Filler", 26, "Spaces."),
        ("82-116", "Narration", 35, "Left justified, padded on the right with spaces."),
        ("117-186", "Filler", 70, "Spaces."),
    ]
    for i, (cols, name, length, content) in enumerate(fields, start=1):
        r = 4 + i
        ws.cell(row=r, column=2, value=i)
        ws.cell(row=r, column=3, value=cols)
        ws.cell(row=r, column=4, value=name)
        ws.cell(row=r, column=5, value=length)
        ws.cell(row=r, column=6, value=content)
        for col in range(2, 7):
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).font = Font(size=10)
        ws.cell(row=r, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=13, column=4, value="Total").font = HEADER
    ws.cell(row=13, column=5, value=186).font = HEADER
    ws.cell(row=13, column=6, value="Records are separated by CR LF.").font = NOTE

    ws.cell(row=15, column=2,
            value="Worked example - the bank's own file for value date 15-07-2026").font = SECTION
    ws.cell(row=16, column=2,
            value="Amounts below are that day's figures; each line is exactly 186 characters.").font = NOTE

    sample = os.path.join(HERE, "..", "samples", "PROPELG_TTUM_16072026_Revised.txt")
    with open(sample, "rb") as handle:
        lines = handle.read().decode("ascii").split("\r\n")
    for i, line in enumerate(lines):
        c = ws.cell(row=17 + i, column=2, value=line)
        c.font = MONO
    ws.sheet_view.showGridLines = False


IMPORT_FIRST, IMPORT_LAST = 10, 109


def build_import(ws):
    """Where an incoming settlement file lands, and what it was read as.

    The macro writes the file's lines into column B; the columns beside it pull the
    fields out by formula, so the same view works when the lines are pasted in by
    hand because macros are blocked.
    """
    banner(ws, "IMPORT FROM SETTLEMENT FILE",
           "The day's figures, read out of the file the settlement system produces.", 9)
    for col, w in {"A": 5, "B": 62, "C": 16, "D": 13, "E": 16, "F": 7,
                   "G": 34, "H": 16, "I": 26, "J": 5}.items():
        ws.column_dimensions[col].width = w

    lines = [
        ("With macros working:", "click Import Latest Input File on the Dashboard. The file's "
         "lines land here, you get a summary to check, and the amounts go onto Entries."),
        ("With macros blocked:", "open the settlement file in Notepad, select all, copy, and "
         "paste into cell B10 below. Then copy the Amount to key column onto the Amount "
         "column of the Entries sheet, matching the rows named under Goes to."),
    ]
    for i, (head, rest) in enumerate(lines):
        r = 3 + i
        ws.cell(row=r, column=2, value=head).font = Font(name="Calibri", size=10, bold=True,
                                                         color="1F3355")
        c = ws.cell(row=r, column=3, value=rest)
        c.font = Font(name="Calibri", size=10, color="1F2937")
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)

    ws.cell(row=6, column=2, value="Last import").font = SECTION
    ws.merge_cells("C6:I6")
    ws["C6"] = "Nothing imported yet."
    ws["C6"].font = Font(name="Calibri", size=10, italic=True, color="6B7280")

    ws.merge_cells("B7:I7")
    ws["B7"] = ('="Lines below  "&COUNTIF($B$%d:$B$%d,"?*")'
                '&"      Debit  "&TEXT(SUMIF($F$%d:$F$%d,"D",$E$%d:$E$%d)/100,"#,##0.00")'
                '&"      Credit  "&TEXT(SUMIF($F$%d:$F$%d,"C",$E$%d:$E$%d)/100,"#,##0.00")'
                '&"      Difference  "&TEXT((SUMIF($F$%d:$F$%d,"D",$E$%d:$E$%d)'
                '-SUMIF($F$%d:$F$%d,"C",$E$%d:$E$%d))/100,"#,##0.00")') % ((
                    IMPORT_FIRST, IMPORT_LAST) * 9)
    ws["B7"].font = Font(name="Calibri", size=11, bold=True)
    ws["B7"].alignment = Alignment(vertical="center")
    ws.row_dimensions[7].height = 20

    headers = ["#", "Line from the settlement file", "Account", "Value date",
               "Amount (paise)", "Dr/Cr", "Narration", "Amount to key", "Goes to"]
    for i, text in enumerate(headers, start=1):
        c = ws.cell(row=IMPORT_FIRST - 1, column=i, value=text)
        c.font = HEADER
        c.fill = HEAD_FILL
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[IMPORT_FIRST - 1].height = 26

    ef, el = FIRST_ROW, LAST_ROW
    for r in range(IMPORT_FIRST, IMPORT_LAST + 1):
        i = r - IMPORT_FIRST + 1
        ws.cell(row=r, column=1, value=i).font = Font(name="Calibri", size=9, color="9AA5B1")
        b = ws.cell(row=r, column=2)
        b.font = MONO
        b.fill = INPUT_FILL
        b.number_format = "@"
        ws.cell(row=r, column=3, value='=IF(LEN($B%d)<14,"",TRIM(MID($B%d,1,14)))' % (r, r))
        ws.cell(row=r, column=4,
                value='=IF(LEN($B%d)<22,"",DATE(VALUE(MID($B%d,19,4)),'
                      'VALUE(MID($B%d,17,2)),VALUE(MID($B%d,15,2))))' % (r, r, r, r))
        ws.cell(row=r, column=4).number_format = "dd-mmm-yyyy"
        ws.cell(row=r, column=5, value='=IF(LEN($B%d)<45,"",VALUE(MID($B%d,23,23)))' % (r, r))
        ws.cell(row=r, column=5).number_format = "#,##0"
        ws.cell(row=r, column=6, value='=IF(LEN($B%d)<46,"",MID($B%d,46,1))' % (r, r))
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=7, value='=IF(LEN($B%d)<116,"",TRIM(MID($B%d,82,35)))' % (r, r))
        ws.cell(row=r, column=8,
                value='=IF($E%d="","",$E%d/IF(ttAmountUnit="Paise",1,100))' % (r, r))
        ws.cell(row=r, column=8).number_format = "#,##0.00"
        ws.cell(row=r, column=8).font = Font(name="Calibri", size=10, bold=True, color="1F3355")
        # Which Entries row claims this line: the one whose match text the
        # narration contains. SUMPRODUCT keeps it a normal formula, not an array one.
        ws.cell(row=r, column=10,
                value='=IF($G%d="","",SUMPRODUCT(MAX((Entries!$H$%d:$H$%d<>"")'
                      '*ISNUMBER(SEARCH(Entries!$H$%d:$H$%d,$G%d&""))'
                      '*(ROW(Entries!$H$%d:$H$%d)-%d))))'
                      % (r, ef, el, ef, el, r, ef, el, ef - 1))
        ws.cell(row=r, column=10).font = Font(name="Calibri", size=8, color="C7CDD4")
        ws.cell(row=r, column=9,
                value='=IF($G%d="","",IF($J%d=0,"-- no match --",'
                      'INDEX(Entries!$B$%d:$B$%d,$J%d)))' % (r, r, ef, el, r))
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = BOX
            if col not in (2, 8):
                ws.cell(row=r, column=col).font = Font(name="Calibri", size=10)

    ws.freeze_panes = "A%d" % IMPORT_FIRST
    ws.sheet_view.showGridLines = False


RECORD_FORMULA = (
    '=IF($C{r}="","",'
    'LEFT(INDEX(Entries!$C${f}:$C${l},$C{r})&REPT(" ",14),14)'
    '&TEXT(ttValueDate,"ddmmyyyy")'
    '&TEXT(INDEX(Entries!$E${f}:$E${l},$C{r})*IF(ttAmountUnit="Paise",1,100),REPT("0",23))'
    '&INDEX(Entries!$D${f}:$D${l},$C{r})'
    '&"0"&TEXT(ttValueDate,"ddmmyyyy")'
    '&REPT(" ",26)'
    '&LEFT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE('
    'INDEX(Entries!$F${f}:$F${l},$C{r}),'
    '"{{DDMMYYYY}}",TEXT(ttValueDate,"ddmmyyyy")),'
    '"{{DDMMYY}}",TEXT(ttValueDate,"ddmmyy")),'
    '"{{YYYY}}",TEXT(ttValueDate,"yyyy")),'
    '"{{DD}}",TEXT(ttValueDate,"dd")),'
    '"{{MM}}",TEXT(ttValueDate,"mm")),'
    '"{{YY}}",TEXT(ttValueDate,"yy"))&REPT(" ",35),35)'
    '&REPT(" ",70))'
)


def build_textoutput(ws):
    """A macro-free route to the same file: the records, built by formula.

    Everything the macro does to lay a record out is done here in worksheet
    functions, so the day's file can still be produced when macros are blocked.
    """
    banner(ws, "TEXT OUTPUT  (works without macros)",
           "The same records, built by formula. Copy them into Notepad and save.", 3)
    for col, w in {"A": 118, "B": 9, "C": 7}.items():
        ws.column_dimensions[col].width = w

    steps = [
        "1.  Set the value date and the amounts as usual, on the Dashboard and Entries sheets.",
        "2.  Select the filled cells in column A below, from A6 down to the last one with text.",
        "3.  Copy, then paste into Notepad.",
        "4.  Save from Notepad as the file name shown on the Dashboard, with Encoding set to ANSI.",
    ]
    for i, text in enumerate(steps):
        c = ws.cell(row=3 + i, column=1, value=text)
        c.font = Font(name="Calibri", size=10, color="1F2937")
    ws.cell(row=3, column=1).font = Font(name="Calibri", size=10, bold=True, color="1F2937")

    ws.cell(row=8, column=1, value="Record  (each line is exactly 186 characters)").font = HEADER
    ws.cell(row=8, column=2, value="Length").font = HEADER
    ws.cell(row=8, column=3, value="Row").font = HEADER
    for col in (1, 2, 3):
        ws.cell(row=8, column=col).fill = HEAD_FILL
        ws.cell(row=8, column=col).border = BOX

    first = 9
    for i in range(LAST_ROW - FIRST_ROW + 1):
        r = first + i
        ws.cell(row=r, column=3,
                value='=IFERROR(MATCH(%d,Entries!$I$%d:$I$%d,0),"")'
                      % (i + 1, FIRST_ROW, LAST_ROW))
        ws.cell(row=r, column=3).font = Font(name="Calibri", size=9, color="9AA5B1")
        ws.cell(row=r, column=1,
                value=RECORD_FORMULA.format(r=r, f=FIRST_ROW, l=LAST_ROW))
        ws.cell(row=r, column=1).font = MONO
        ws.cell(row=r, column=2, value='=IF($A%d="","",LEN($A%d))' % (r, r))
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=9, color="6B7280")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A%d" % first
    ws.sheet_view.showGridLines = False


def build_setup(ws):
    """What to try when the buttons do not respond."""
    banner(ws, "SETUP AND TROUBLESHOOTING",
           "Read this if clicking a button on the Dashboard does nothing.", 3)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 104

    blocks = [
        ("Why a button can do nothing", None),
        ("The buttons are drawings stored in the file, so they are always visible. What they "
         "run is a macro, and Excel will refuse to run macros in a file it does not trust. "
         "Work down this list - the first step fixes it most of the time.", "note"),

        ("1.  Unblock the file  (do this one first)", "head"),
        ("Close the workbook. Find it in File Explorer, right-click it, choose Properties, tick "
         "Unblock at the bottom of the General tab, click OK, then open it again.", None),
        ("Files that arrive by email or download carry a mark that makes Excel block their "
         "macros outright. Newer Excel shows a red bar for this - Enable Content will not appear, "
         "and clicking anything else will not help. Unblocking is the only fix.", "note"),

        ("2.  Put the workbook in a trusted folder", "head"),
        ("File > Options > Trust Center > Trust Center Settings > Trusted Locations > "
         "Add new location. Add the folder you keep this workbook in, then reopen it.", None),

        ("3.  Check the macro setting itself", "head"),
        ("File > Options > Trust Center > Trust Center Settings > Macro Settings. It should be "
         "Disable all macros with notification. If it is Disable all macros without "
         "notification, no prompt ever appears and nothing will run.", None),

        ("4.  Check whether the macros arrived at all", "head"),
        ("Press Alt+F8. If GenerateTTUM is in the list, the macros are present - run it straight "
         "from there, and the buttons will work once trust is sorted out. If the list is empty, "
         "the macros did not load, so do step 5.", None),

        ("5.  Load the macros by hand  (about a minute, needed only once)", "head"),
        ("Press Alt+F11 to open the Visual Basic editor. Choose File > Import File, pick "
         "modTTUM.bas from the folder this workbook came in, then press Alt+Q to close the "
         "editor. Save the workbook. The buttons are already wired to it and will now work.", None),

        ("If macros are blocked and cannot be unblocked", "head"),
        ("Use the Text Output sheet. It builds the same records with worksheet formulas and no "
         "macro at all - copy them into Notepad and save. Nothing else about the workbook "
         "changes: the dates, amounts and checks all work as normal.", None),
    ]

    row = 4
    for text, kind in blocks:
        c = ws.cell(row=row, column=2, value=text)
        if kind == "head":
            c.font = SECTION
            c.fill = BAND
            row += 1
        elif kind == "note":
            c.font = NOTE
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 42
            row += 2
        else:
            c.font = Font(name="Calibri", size=11, color="1F2937") if text.startswith("Why") \
                else LABEL
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 44
            row += 2
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------------- buttons

EMU_PER_POINT = 12700

# label, macro, and whether it is the primary action
BUTTONS = [
    ("Generate TTUM File", "GenerateTTUM", True),
    ("Import Latest Input File", "ImportLatestFile", True),
    ("Choose Input File...", "ChooseInputFile", False),
    ("Preview Records", "PreviewTTUM", False),
    ("Validate Entries", "ValidateEntries", False),
    ("Reset Date to Today", "ResetDateToToday", False),
    ("Clear Amounts", "ClearAmounts", False),
    ("Choose Output Folder...", "BrowseOutputFolder", False),
    ("Open Output Folder", "OpenOutputFolder", False),
]

BUTTON_COL = 7          # column H, zero based
BUTTON_ROW = 4          # row 5, zero based
BUTTON_WIDTH = 160      # points
BUTTON_HEIGHT = 30
BUTTON_PITCH = 38


def _button_shape(index, label, macro, primary):
    fill = "1F3355" if primary else "F1F5FA"
    line = "16283F" if primary else "9FB2C6"
    text = "FFFFFF" if primary else "1F3355"
    return (
        '<xdr:oneCellAnchor>'
        '<xdr:from><xdr:col>%d</xdr:col><xdr:colOff>0</xdr:colOff>'
        '<xdr:row>%d</xdr:row><xdr:rowOff>%d</xdr:rowOff></xdr:from>'
        '<xdr:ext cx="%d" cy="%d"/>'
        '<xdr:sp macro="[0]!%s" textlink="">'
        '<xdr:nvSpPr><xdr:cNvPr id="%d" name="btnTT%d"/><xdr:cNvSpPr/></xdr:nvSpPr>'
        '<xdr:spPr>'
        '<a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
        '<a:prstGeom prst="roundRect"><a:avLst><a:gd name="adj" fmla="val 12000"/></a:avLst></a:prstGeom>'
        '<a:solidFill><a:srgbClr val="%s"/></a:solidFill>'
        '<a:ln w="12700"><a:solidFill><a:srgbClr val="%s"/></a:solidFill></a:ln>'
        '</xdr:spPr>'
        '<xdr:txBody>'
        '<a:bodyPr vertOverflow="clip" horzOverflow="clip" rtlCol="0" anchor="ctr"/>'
        '<a:lstStyle/>'
        '<a:p><a:pPr algn="ctr"/>'
        '<a:r><a:rPr lang="en-US" sz="1000" b="%d"><a:solidFill><a:srgbClr val="%s"/></a:solidFill>'
        '<a:latin typeface="Calibri"/></a:rPr><a:t>%s</a:t></a:r></a:p>'
        '</xdr:txBody>'
        '</xdr:sp>'
        '<xdr:clientData/>'
        '</xdr:oneCellAnchor>'
    ) % (BUTTON_COL, BUTTON_ROW, index * BUTTON_PITCH * EMU_PER_POINT,
         BUTTON_WIDTH * EMU_PER_POINT, BUTTON_HEIGHT * EMU_PER_POINT,
         macro, index + 2, index, fill, line, 1 if primary else 0, text, label)


def build_drawing_xml():
    """The Dashboard buttons, stored in the file rather than drawn at run time.

    Shapes carrying a `macro` attribute are ordinary drawings, so they are present
    and visible even before macros are enabled - which is the difference between
    a workbook that looks broken and one that just needs Enable Content.
    """
    shapes = "".join(_button_shape(i, label, macro, primary)
                     for i, (label, macro, primary) in enumerate(BUTTONS))
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/'
            'spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
            + shapes + '</xdr:wsDr>').encode("utf-8")


def attach_drawing(name, data, dashboard_part):
    """Wires the drawing into the Dashboard worksheet part."""
    if name == dashboard_part:
        text = data.decode("utf-8")
        if "xmlns:r=" not in text:
            text = text.replace(
                "<worksheet ",
                '<worksheet xmlns:r="http://schemas.openxmlformats.org/'
                'officeDocument/2006/relationships" ', 1)
        text = text.replace("</worksheet>", '<drawing r:id="rIdDrawing"/></worksheet>')
        return text.encode("utf-8")
    return data


# ------------------------------------------------------------------- build

SHEETS = [
    ("Dashboard", "shDashboard", build_dashboard),
    ("Entries", "shEntries", build_entries),
    ("Import", "shImport", build_import),
    ("Text Output", "shTextOutput", build_textoutput),
    ("Config", "shConfig", build_config),
    ("Log", "shLog", build_log),
    ("Layout", "shLayout", build_layout),
    ("Setup", "shSetup", build_setup),
]

NAMES = {
    "ttValueDate": "Dashboard!$C$6",
    "ttFileDate": "Dashboard!$C$7",
    "ttOutputFolder": "Dashboard!$C$17",
    "ttStatus": "Dashboard!$B$22",
    "ttAmountUnit": "Config!$C$5",
    "ttFileNamePattern": "Config!$C$6",
    "ttFileDateOffset": "Config!$C$7",
    "ttEnforceBalance": "Config!$C$8",
    "ttTrailingNewline": "Config!$C$9",
    "ttInputFolder": "Config!$C$10",
    "ttInputPattern": "Config!$C$11",
    "ttImportSetsDate": "Config!$C$12",
    "ttLastImport": "Import!$C$6",
}


def build_workbook_xlsx(path):
    wb = Workbook()
    wb.remove(wb.active)
    wb.code_name = "ThisWorkbook"
    for title, code_name, builder in SHEETS:
        ws = wb.create_sheet(title)
        ws.sheet_properties.codeName = code_name
        builder(ws)
    for name, ref in NAMES.items():
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    wb.active = 0
    wb.save(path)


def read_source(name):
    with open(os.path.join(VBA, name), "r", encoding="utf-8") as handle:
        return handle.read()


def build_vba():
    """The VBA project: one document module per sheet, plus ThisWorkbook and modTTUM."""
    body = read_source("ThisWorkbook.cls").split("Attribute VB_Exposed = True", 1)[1]
    modules = [vbaproject.Module(
        "ThisWorkbook",
        vbaproject.document_header("ThisWorkbook", True) + body,
        vbaproject.MODULE_DOCUMENT)]
    for title, code_name, _ in SHEETS:
        modules.append(vbaproject.Module(
            code_name, vbaproject.document_header(code_name, False),
            vbaproject.MODULE_DOCUMENT))
    modules.append(vbaproject.Module("modTTUM", read_source("modTTUM.bas")))
    return vbaproject.build("VBAProject", modules)


def to_macro_enabled(xlsx_path, xlsm_path, vba_blob):
    """Repackages the .xlsx as a macro-enabled workbook carrying vbaProject.bin."""
    src = zipfile.ZipFile(xlsx_path)
    if os.path.exists(xlsm_path):
        os.remove(xlsm_path)
    out = zipfile.ZipFile(xlsm_path, "w", zipfile.ZIP_DEFLATED)

    dashboard_part = None
    for name in src.namelist():
        if name.startswith("xl/worksheets/sheet") and \
                b'codeName="shDashboard"' in src.read(name):
            dashboard_part = name
            break
    if dashboard_part is None:
        raise RuntimeError("could not find the Dashboard worksheet part")

    for item in src.infolist():
        data = attach_drawing(item.filename, src.read(item.filename), dashboard_part)

        if item.filename == "[Content_Types].xml":
            text = data.decode("utf-8")
            text = text.replace(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                "application/vnd.ms-excel.sheet.macroEnabled.main+xml")
            insert = ('<Default Extension="bin" '
                      'ContentType="application/vnd.ms-office.vbaProject"/>'
                      '<Override PartName="/xl/drawings/drawing1.xml" ContentType='
                      '"application/vnd.openxmlformats-officedocument.drawing+xml"/>')
            text = text.replace("</Types>", insert + "</Types>")
            data = text.encode("utf-8")

        elif item.filename == "xl/workbook.xml":
            text = data.decode("utf-8")
            if "codeName=" not in text:
                if re.search(r"<workbookPr[^>]*/>", text):
                    text = re.sub(r"<workbookPr([^>]*)/>",
                                  r'<workbookPr\1 codeName="ThisWorkbook"/>', text, count=1)
                else:
                    text = text.replace("<sheets>",
                                        '<workbookPr codeName="ThisWorkbook"/><sheets>', 1)
            data = text.encode("utf-8")

        elif item.filename == "xl/_rels/workbook.xml.rels":
            text = data.decode("utf-8")
            rel = ('<Relationship Id="rIdVBA" '
                   'Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" '
                   'Target="vbaProject.bin"/>')
            text = text.replace("</Relationships>", rel + "</Relationships>")
            data = text.encode("utf-8")

        out.writestr(item, data)

    out.writestr("xl/vbaProject.bin", vba_blob)
    out.writestr("xl/drawings/drawing1.xml", build_drawing_xml())
    out.writestr(
        "xl/worksheets/_rels/%s.rels" % os.path.basename(dashboard_part),
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rIdDrawing" Type="http://schemas.openxmlformats.org/'
        'officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
        '</Relationships>')
    out.close()
    src.close()


def main():
    if not os.path.isdir(DIST):
        os.makedirs(DIST)
    xlsx = os.path.join(DIST, "TTUM_Generator_NoMacros.xlsx")
    xlsm = os.path.join(DIST, "TTUM_Generator.xlsm")

    build_workbook_xlsx(xlsx)
    to_macro_enabled(xlsx, xlsm, build_vba())
    shutil.copyfile(os.path.join(VBA, "modTTUM.bas"), os.path.join(DIST, "modTTUM.bas"))

    for path in (xlsm, xlsx):
        print("%-42s %8d bytes" % (os.path.basename(path), os.path.getsize(path)))


if __name__ == "__main__":
    main()
